from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.services.rules import RuleEngine


HEADER_ALIASES = {
    'employeeNo': ['工号', '员工工号', 'employee_no', 'employeeNo'],
    'name': ['姓名', '员工姓名', 'name'],
    'department': ['部门', '所属部门', 'department'],
    'date': ['考勤日期', '日期', '打卡日期', 'date'],
    'shouldAttendance': ['应出勤', '应出勤天数', 'should_attendance', 'shouldAttendance'],
    'actualAttendance': ['实出勤', '实出勤天数', 'actual_attendance', 'actualAttendance'],
    'checkIn': ['上班打卡时间', '上班打卡', '签到', 'check_in', 'checkIn'],
    'checkOut': ['下班打卡时间', '下班打卡', '签退', 'check_out', 'checkOut'],
    'status': ['考勤状态', '状态', 'attendance_status'],
    'lateMinutes': ['迟到分钟', 'late_minutes', 'lateMinutes'],
    'earlyMinutes': ['早退分钟', 'early_minutes', 'earlyMinutes'],
    'absenteeismDays': ['旷工天数', '旷工', 'absenteeism_days', 'absenteeismDays'],
    'missingCardCount': ['缺卡次数', '缺卡', 'missing_card_count', 'missingCardCount'],
    'leaveType': ['请假类型', 'leave_type', 'leaveType'],
    'overtimeHours': ['加班小时', 'overtime_hours', 'overtimeHours'],
    'remark': ['备注', 'remark'],
}


class ExcelAttendanceService:
    def __init__(self) -> None:
        self.rule_engine = RuleEngine()

    @staticmethod
    def _as_text(value: Any) -> str:
        return '' if value is None else str(value).strip()

    @staticmethod
    def _as_number(value: Any, default: float = 0.0) -> float:
        if value in (None, ''):
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default

    def parse(self, content: bytes, month: str = '') -> dict[str, Any]:
        if not content:
            return {'records': [], 'errors': [{'row': '', 'message': '上传文件为空'}], 'matched_fields': {}, 'total_count': 0, 'success_count': 0, 'error_count': 1}
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return {'records': [], 'errors': ['Excel内容为空'], 'matched_fields': {}}

        headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
        field_mapping = self.match_fields(headers)
        required_fields = ['employeeNo', 'name', 'department', 'date', 'checkIn', 'checkOut', 'status', 'lateMinutes', 'earlyMinutes', 'leaveType', 'overtimeHours', 'remark']
        missing_fields = [field for field in required_fields if field not in field_mapping]
        if missing_fields:
            return {'records': [], 'errors': [f"缺少字段: {', '.join(missing_fields)}"], 'matched_fields': field_mapping}
        records = []
        errors = []

        for index, row in enumerate(rows[1:], start=2):
            payload = {}
            for field, header in field_mapping.items():
                header_index = headers.index(header)
                payload[field] = row[header_index] if header_index < len(row) else None
            if not payload.get('employeeNo') or not payload.get('date'):
                errors.append({'row': index, 'employeeNo': self._as_text(payload.get('employeeNo')), 'name': self._as_text(payload.get('name')), 'department': self._as_text(payload.get('department')), 'message': '缺少工号或考勤日期'})
                continue
            check_in = self._as_text(payload.get('checkIn'))[:5] if payload.get('checkIn') else None
            check_out = self._as_text(payload.get('checkOut'))[:5] if payload.get('checkOut') else None
            judgement = self.rule_engine.evaluate_attendance(check_in, check_out)
            manual_status = self._as_text(payload.get('status'))
            late_minutes = int(self._as_number(payload.get('lateMinutes'), judgement.late_minutes or 0))
            early_minutes = int(self._as_number(payload.get('earlyMinutes'), judgement.early_minutes or 0))
            overtime_hours = self._as_number(payload.get('overtimeHours'), 0)
            leave_type = self._as_text(payload.get('leaveType'))
            remark = self._as_text(payload.get('remark'))
            normalized_date = self._as_text(payload['date'])
            month_value = month or normalized_date[:7]
            status_text = manual_status or judgement.status
            should_attendance = self._as_number(payload.get('shouldAttendance'), 1)
            default_actual_attendance = 0 if any(keyword in status_text for keyword in ['旷工', '请假']) else 1
            actual_attendance = self._as_number(payload.get('actualAttendance'), default_actual_attendance)
            absenteeism_days = self._as_number(payload.get('absenteeismDays'), 1 if '旷工' in status_text else 0)
            missing_card_count = int(self._as_number(payload.get('missingCardCount'), 1 if '缺卡' in status_text else 0))
            records.append({
                '_row': index,
                'employeeNo': self._as_text(payload['employeeNo']),
                'name': self._as_text(payload.get('name')),
                'department': self._as_text(payload.get('department')),
                'date': normalized_date,
                'month': month_value,
                'shouldAttendance': round(should_attendance, 2),
                'actualAttendance': round(actual_attendance, 2),
                'checkIn': self._as_text(payload.get('checkIn')),
                'checkOut': self._as_text(payload.get('checkOut')),
                'status': status_text,
                'lateMinutes': late_minutes,
                'earlyMinutes': early_minutes,
                'absenteeismDays': round(absenteeism_days, 2),
                'missingCardCount': missing_card_count,
                'leaveType': leave_type,
                'overtimeHours': overtime_hours,
                'workHours': round(judgement.work_hours + overtime_hours, 2),
                'remark': remark,
                'type': '工作日',
            })

        return {
            'records': records,
            'errors': errors,
            'matched_fields': field_mapping,
            'total_count': len(records) + len(errors),
            'success_count': len(records),
            'error_count': len(errors),
        }

    def match_fields(self, headers: list[str]) -> dict[str, str]:
        result = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in headers:
                    result[field] = alias
                    break
        return result
