from io import BytesIO

from openpyxl import load_workbook

from app.performance.constants import IMPORT_HEADERS
from app.performance.mappers import (
    calculate_total_score,
    coefficient_from_grade,
    employee_lookup,
    grade_from_score,
    hydrate_performance_record,
    normalize_text,
    to_int,
    to_number,
)
from app.performance.queries import filter_rows
from app.performance.permissions import scope_rows
from app.performance.validators import validate_row


def export_records(rows: list[dict], cycle_type: str = '', assessment_year: int | None = None, assessment_month: int | None = None) -> dict:
    return {
        'fileName': f"绩效报表_{cycle_type or '全部'}_{assessment_year or '全部'}_{assessment_month or '全部'}",
        'records': [
            {
                '工号': item.get('employeeNo', ''),
                '姓名': item.get('name', ''),
                '部门': item.get('department', ''),
                '岗位': item.get('position', ''),
                '绩效周期': item.get('cycleType', ''),
                '考核年份': item.get('assessmentYear', ''),
                '考核月份': item.get('assessmentMonth', ''),
                '业绩指标得分': item.get('performanceScore', 0),
                '工作态度得分': item.get('attitudeScore', 0),
                '能力表现得分': item.get('abilityScore', 0),
                '综合总分': item.get('totalScore', 0),
                '绩效等级': item.get('grade', ''),
                '绩效系数': item.get('coefficient', 0),
                '考核状态': item.get('status', ''),
                '备注': item.get('remark', ''),
            }
            for item in rows
        ],
    }


def match_headers(headers: list[str]) -> dict[str, str]:
    result = {}
    for field, aliases in IMPORT_HEADERS.items():
        for alias in aliases:
            if alias in headers:
                result[field] = alias
                break
    return result


def parse_import(content: bytes, cycle_type: str, assessment_year: int, assessment_month: int | None) -> dict:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    if not rows:
        return {'records': [], 'errors': [{'row': '', 'message': 'Excel内容为空'}], 'matched_fields': {}, 'total_count': 0, 'success_count': 0, 'error_count': 1}
    headers = [str(cell).strip() if cell is not None else '' for cell in rows[0]]
    mapping = match_headers(headers)
    missing = [field for field in IMPORT_HEADERS if field not in mapping and field != 'assessmentMonth']
    if missing:
        return {'records': [], 'errors': [{'row': '', 'message': f"缺少字段: {', '.join(missing)}"}], 'matched_fields': mapping, 'total_count': 0, 'success_count': 0, 'error_count': 1}
    records, errors = [], []
    for row_index, row in enumerate(rows[1:], start=2):
        payload = {}
        for field, header in mapping.items():
            index = headers.index(header)
            payload[field] = row[index] if index < len(row) else None
        employee_no = normalize_text(payload.get('employeeNo'))
        if not employee_no:
            errors.append({'row': row_index, 'employeeNo': '', 'message': '工号不能为空'})
            continue
        records.append({
            'employeeNo': employee_no,
            'name': normalize_text(payload.get('name')),
            'department': normalize_text(payload.get('department')),
            'position': normalize_text(payload.get('position')),
            'cycleType': normalize_text(payload.get('cycleType')) or cycle_type,
            'assessmentYear': to_int(payload.get('assessmentYear'), assessment_year),
            'assessmentMonth': to_int(payload.get('assessmentMonth'), assessment_month),
            'performanceScore': to_number(payload.get('performanceScore')),
            'attitudeScore': to_number(payload.get('attitudeScore')),
            'abilityScore': to_number(payload.get('abilityScore')),
            'totalScore': to_number(payload.get('totalScore')),
            'grade': normalize_text(payload.get('grade')),
            'coefficient': to_number(payload.get('coefficient')),
            'selfReview': normalize_text(payload.get('selfReview')),
            'managerReview': normalize_text(payload.get('managerReview')),
            'status': normalize_text(payload.get('status')) or '待自评',
            'remark': normalize_text(payload.get('remark')),
            'indicators': [],
        })
    return {'records': records, 'errors': errors, 'matched_fields': mapping, 'total_count': len(records) + len(errors), 'success_count': len(records), 'error_count': len(errors)}


def parse_import_result(content: bytes, cycle_type: str, assessment_year: int, assessment_month: int | None, repository) -> dict:
    employee_map = employee_lookup(repository)
    parsed = parse_import(content, cycle_type, assessment_year, assessment_month)
    rows, errors = [], list(parsed['errors'])
    for item in parsed['records']:
        row = hydrate_performance_record(item, employee_map)
        if row.get('employeeNo') not in employee_map:
            errors.append({'row': '', 'employeeNo': row.get('employeeNo', ''), 'message': '工号不存在，无法关联员工档案'})
            continue
        if to_number(row.get('totalScore')) == 0 and any([row.get('performanceScore'), row.get('attitudeScore'), row.get('abilityScore')]):
            row['totalScore'] = calculate_total_score(row)
            row['score'] = row['totalScore']
            row['grade'] = row.get('grade') or grade_from_score(row['totalScore'])
            row['coefficient'] = row.get('coefficient') or coefficient_from_grade(row['grade'])
        rows.append(row)
    parsed.update({'records': rows, 'errors': errors, 'success_count': len(rows), 'error_count': len(errors), 'total_count': len(rows) + len(errors)})
    return parsed


def confirm_import_records(payload: dict, repository) -> list[dict]:
    employee_map = employee_lookup(repository)
    stored = []
    for item in payload.get('records') or []:
        row = hydrate_performance_record(item, employee_map)
        if not row.get('employeeNo'):
            continue
        validate_row(row)
        stored.append(repository.upsert('performance', row))
    repository.upsert('performance_imports', {
        'cycleType': payload.get('cycleType', '月度'),
        'assessmentYear': payload.get('assessmentYear'),
        'assessmentMonth': payload.get('assessmentMonth'),
        'importedCount': len(stored),
        'errors': payload.get('errors') or [],
        'status': '已导入',
    })
    return stored


def export_dataset(repository, user: dict, actor: dict, keyword: str = '', department: str = '', position: str = '', cycle_type: str = '', assessment_year: int | None = None, assessment_month: int | None = None, status: str = '') -> dict:
    employee_map = employee_lookup(repository)
    rows = [hydrate_performance_record(item, employee_map) for item in repository.list('performance')]
    rows = filter_rows(scope_rows(rows, user, actor), keyword, department, position, cycle_type, assessment_year, assessment_month, status)
    return export_records(rows, cycle_type, assessment_year, assessment_month)
